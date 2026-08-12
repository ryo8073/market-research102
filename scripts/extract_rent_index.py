"""
マンション賃料インデックス Excel → JSON 変換スクリプト

三井住友トラスト基礎研究所 × アットホーム
MCI_DDL.xlsx（無償版Excel）から12エリア×4タイプの全時系列データを抽出し、
ci102-nextjs/public/data/rent_index.json として出力する。

Usage:
    python scripts/extract_rent_index.py
    python scripts/extract_rent_index.py --xlsx data/cache/MCI_DDL.xlsx
"""
import io
import sys
import json
import os
import argparse

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl

# エリア名 → キー のマッピング
AREA_MAP = {
    "東京23区":     {"key": "tokyo_23ku",        "pref_code": 13},
    "東京都下":     {"key": "tokyo_tama",         "pref_code": 13},
    "横浜・川崎市": {"key": "yokohama_kawasaki",   "pref_code": 14},
    "千葉西部":     {"key": "chiba_west",         "pref_code": 12},
    "埼玉東南部":   {"key": "saitama_se",          "pref_code": 11},
    "札幌市":       {"key": "sapporo",            "pref_code": 1},
    "仙台市":       {"key": "sendai",             "pref_code": 4},
    "名古屋市":     {"key": "nagoya",             "pref_code": 23},
    "京都市":       {"key": "kyoto",              "pref_code": 26},
    "大阪市":       {"key": "osaka",              "pref_code": 27},
    "大阪広域":     {"key": "osaka_wide",         "pref_code": 27},
    "福岡市":       {"key": "fukuoka",            "pref_code": 40},
}

TYPE_MAP = {
    "シングル": "single",
    "コンパクト": "compact",
    "ファミリー": "family",
    "総合": "total",
}


def parse_sheet(ws) -> dict | None:
    """1シートからエリアデータを抽出"""
    area_name = ws.title.strip()
    if area_name not in AREA_MAP:
        return None

    meta = AREA_MAP[area_name]

    # ヘッダー行を探す（シングル/コンパクト/ファミリー/総合）
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        vals = [str(ws.cell(row=row_idx, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if "シングル" in vals and "総合" in vals:
            header_row = row_idx
            break

    if header_row is None:
        print(f"  Warning: {area_name} - header not found", file=sys.stderr)
        return None

    # カラムインデックスを特定
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "").strip()
        for jp, en in TYPE_MAP.items():
            if val == jp:
                col_map[en] = c
                break

    # 年四半期列を特定（"2009.Q1"等を含む列）
    period_col = None
    for c in range(1, ws.max_column + 1):
        for r in range(header_row + 1, min(header_row + 3, ws.max_row + 1)):
            val = str(ws.cell(row=r, column=c).value or "")
            if "2009" in val and "Q1" in val:
                period_col = c
                break
        if period_col:
            break

    if not period_col or not col_map:
        print(f"  Warning: {area_name} - columns not found", file=sys.stderr)
        return None

    # データ行を読む
    timeseries = {}  # type_key -> [(period, value), ...]
    for type_key in col_map:
        timeseries[type_key] = []

    for r in range(header_row + 1, ws.max_row + 1):
        period = str(ws.cell(row=r, column=period_col).value or "").strip()
        if not period or "Q" not in period:
            continue
        # "2009.Q1" → "2009Q1"
        period_clean = period.replace(".", "")

        for type_key, col in col_map.items():
            val = ws.cell(row=r, column=col).value
            if val is not None:
                try:
                    timeseries[type_key].append({
                        "period": period_clean,
                        "value": round(float(val), 2),
                    })
                except (ValueError, TypeError):
                    pass

    n_periods = max(len(v) for v in timeseries.values()) if timeseries else 0
    print(f"  {area_name}: {len(col_map)} types x {n_periods} quarters")

    return {
        "key": meta["key"],
        "label": area_name,
        "pref_code": meta["pref_code"],
        "timeseries": timeseries,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="data/cache/MCI_DDL.xlsx")
    parser.add_argument("--output", default="ci102-nextjs/public/data/rent_index.json")
    args = parser.parse_args()

    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx_path = os.path.join(base, args.xlsx)
    out_path = os.path.join(base, args.output)

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    areas = []
    for name in wb.sheetnames:
        if name in AREA_MAP:
            result = parse_sheet(wb[name])
            if result:
                areas.append(result)

    wb.close()

    # 期間を特定
    all_periods = set()
    for area in areas:
        for ts in area["timeseries"].values():
            for pt in ts:
                all_periods.add(pt["period"])
    periods_sorted = sorted(all_periods)
    first = periods_sorted[0] if periods_sorted else "?"
    last = periods_sorted[-1] if periods_sorted else "?"

    result = {
        "source": "三井住友トラスト基礎研究所 × アットホーム マンション賃料インデックス",
        "period_range": f"{first}-{last}",
        "latest_period": last,
        "base": "2009Q1=100（連鎖型インデックス）",
        "types_legend": {
            "single": "シングル（18-30m²）",
            "compact": "コンパクト（30-60m²）",
            "family": "ファミリー（60-100m²）",
            "total": "総合",
        },
        "areas": areas,
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\nOutput: {out_path}")
    print(f"Areas: {len(areas)}, Periods: {first} - {last} ({len(periods_sorted)} quarters)")


if __name__ == "__main__":
    main()
